"""Audit trail — logs all create/update/delete actions with user + timestamp to an Excel sheet."""
import logging
import os
from datetime import datetime
from database import excel_db
from config import CONFIG_DIR
from utils.auth import auth_manager

logger = logging.getLogger(__name__)

AUDIT_SHEET = "AuditLog"


def _get_or_create_audit_sheet(wb):
    if AUDIT_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(AUDIT_SHEET)
        ws.append(["Timestamp", "User", "Action", "Entity", "Record_ID", "Details"])
        wb.save()
    return wb[AUDIT_SHEET]


def log(action, entity, record_id, details=""):
    """Log an action to the audit sheet of the current workbook."""
    try:
        if not excel_db._active_file or not os.path.exists(excel_db._active_file):
            logger.warning(f"Audit skipped (no workbook): {action} {entity} #{record_id}")
            return
        from openpyxl import load_workbook
        wb = load_workbook(excel_db._active_file)
        ws = _get_or_create_audit_sheet(wb)
        user = auth_manager.get_current_user() or "system"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, user, action, entity, str(record_id), str(details)])
        wb.save()
        wb.close()
        logger.debug(f"Audit: {user} {action} {entity} #{record_id}")
    except Exception as exc:
        logger.error(f"Audit log failed: {exc}")


def log_create(entity, record_id, details=""):
    log("CREATE", entity, record_id, details)


def log_update(entity, record_id, details=""):
    log("UPDATE", entity, record_id, details)


def log_delete(entity, record_id, details=""):
    log("DELETE", entity, record_id, details)


def log_login(user, success=True):
    try:
        if not excel_db._active_file or not os.path.exists(excel_db._active_file):
            return
        from openpyxl import load_workbook
        wb = load_workbook(excel_db._active_file)
        ws = _get_or_create_audit_sheet(wb)
        status = "LOGIN_OK" if success else "LOGIN_FAIL"
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), user, status, "Auth", "", ""])
        wb.save()
        wb.close()
    except Exception:
        pass