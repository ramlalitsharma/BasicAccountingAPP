import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import json
import shutil
import threading
import functools
import logging
from datetime import datetime
from config import DATA_DIR, BACKUP_DIR, set_setting
from utils.update_checker import get_update_status


logger = logging.getLogger(__name__)
_WB_LOCK = threading.Lock()
_active_file = None
SHEET_VERSION = "2.0"


def _num(v, default=0.0):
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def synchronized(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        with _WB_LOCK:
            return func(*args, **kwargs)
    return wrapper


def _check_write_lock():
    status = get_update_status()
    if status.get("mandatory", False):
        raise PermissionError(
            f"Update Required: v{status['latest_version']} is now available.\n"
            "A mandatory update is pending. Please restart the application to update."
        )

SHEETS = {
    "Suppliers": ["ID", "Name", "Contact", "Address", "Created_At"],
    "Stock": ["ID", "Item_Name", "Category", "Quantity", "Min_Quantity",
              "Purchase_Price", "Selling_Price", "Supplier_ID", "Created_At"],
    "Sales": ["ID", "Stock_ID", "Customer_ID", "Quantity_Sold", "Price", "Total", 
              "Payment_Status", "Paid_Amount", "Unpaid_Amount", "Payment_Date", 
              "Sale_Date", "Receipt_No"],
    "StockLog": ["ID", "Stock_ID", "Change_Type", "Qty_Change", "Old_Qty",
                 "New_Qty", "Note", "Created_At"],
    "Customers": ["ID", "Name", "Contact", "Address", "Created_At"],
    "Purchases": ["ID", "Stock_ID", "Supplier_ID", "Quantity", "Cost_Price",
                  "Total_Cost", "Purchase_Date"],
    "ExtraIncome": ["ID", "Source", "Description", "Amount", "Category", 
                    "Payment_Method", "Reference_No", "Created_At"],
    "Preorders": ["ID", "Customer_ID", "Stock_ID", "Quantity", "Preorder_Price",
                  "Total", "Delivery_Date", "Notes", "Status",
                  "Created_At", "Completed_At", "Sale_ID",
                  "Delivery_Address", "Advance_Amount", "Advance_Payment_Type", "Advance_Paid_At"],
}

INVOICE_PREFIX = "INV"


def set_active_file(filepath):
    global _active_file
    _active_file = filepath
    try:
        set_setting("last_file", filepath)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass


def get_active_file():
    return _active_file


def _backup_file(filepath):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(filepath)
    backup_path = os.path.join(BACKUP_DIR, f"pre_{ts}_{base}")
    try:
        shutil.copy2(filepath, backup_path)
    except (FileNotFoundError, PermissionError, OSError) as e:
        logger.warning("Backup failed for %s: %s", filepath, e)


def _ensure_sheets(wb):
    changed = False
    for name, headers in SHEETS.items():
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
            changed = True
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        changed = True
    return changed


def _validate_sheets(wb):
    issues = []
    for name, headers in SHEETS.items():
        if name in wb.sheetnames:
            ws = wb[name]
            actual = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
            if not actual:
                ws.append(headers)
                issues.append(f"Added missing headers to {name}")
            else:
                for expected in headers:
                    if expected not in actual:
                        missing_idx = headers.index(expected)
                        col_letter = get_column_letter(missing_idx + 1)
                        ws[f"{col_letter}1"] = expected
                        issues.append(f"Fixed missing column '{expected}' in {name}")
    return issues


def create_new_workbook(filepath=None):
    if filepath is None:
        os.makedirs(DATA_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(DATA_DIR, f"accounts_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    ws.append(SHEETS["Suppliers"])
    ws = wb.create_sheet("Stock")
    ws.append(SHEETS["Stock"])
    ws = wb.create_sheet("Sales")
    ws.append(SHEETS["Sales"])
    ws = wb.create_sheet("StockLog")
    ws.append(SHEETS["StockLog"])
    ws = wb.create_sheet("Customers")
    ws.append(SHEETS["Customers"])
    ws = wb.create_sheet("Purchases")
    ws.append(SHEETS["Purchases"])
    ws = wb.create_sheet("ExtraIncome")
    ws.append(SHEETS["ExtraIncome"])
    ws = wb.create_sheet("Preorders")
    ws.append(SHEETS["Preorders"])
    wb.save(filepath)
    set_active_file(filepath)
    return filepath


def open_workbook(filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found:\n{filepath}")
    if not filepath.endswith(".xlsx"):
        raise ValueError("Please select a valid .xlsx file")
    _backup_file(filepath)
    wb = openpyxl.load_workbook(filepath)
    changed = _ensure_sheets(wb)
    issues = _validate_sheets(wb)
    if changed or issues:
        wb.save(filepath)
    set_active_file(filepath)
    return filepath


def _get_wb():
    if not _active_file or not os.path.exists(_active_file):
        raise FileNotFoundError("No active workbook. Create or open one first.")
    return openpyxl.load_workbook(_active_file)


def _save_and_close(wb):
    if _active_file is None:
        wb.close()
        return
    tmp = _active_file + ".tmp"
    try:
        wb.save(tmp)
        shutil.move(tmp, _active_file)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise
    finally:
        wb.close()


def _close_wb(wb):
    try:
        wb.close()
    except Exception:
        pass


def _sheet_to_dicts(ws):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows


def _dicts_to_sheet(ws, dicts, headers):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    ws.delete_rows(2, ws.max_row - 1)
    for d in dicts:
        ws.append([d.get(h, "") for h in headers])


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")



def _next_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (ValueError, TypeError):
                pass
    return max_id + 1


# ---- SUPPLIERS ----

@synchronized
def add_supplier(name, contact="", address=""):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Suppliers"]
        sid = _next_id(ws)
        ws.append([sid, name, contact, address, _now()])
        _save_and_close(wb)
        return sid
    except Exception:
        _close_wb(wb)
        raise


def get_suppliers(search=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Suppliers"])
    wb.close()
    if search:
        search = search.lower()
        data = [r for r in data if search in str(r.get("Name", "")).lower()]
    return data


def get_supplier(supplier_id):
    wb = _get_wb()
    for r in _sheet_to_dicts(wb["Suppliers"]):
        if r.get("ID") == supplier_id:
            wb.close()
            return r
    wb.close()
    return None


@synchronized
def update_supplier(supplier_id, name, contact, address):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Suppliers"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == supplier_id:
                r["Name"] = name
                r["Contact"] = contact
                r["Address"] = address
                break
        _dicts_to_sheet(ws, data, SHEETS["Suppliers"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_supplier(supplier_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        stock_sheet = wb["Stock"] if "Stock" in wb.sheetnames else None
        stock_data = _sheet_to_dicts(stock_sheet) if stock_sheet else []
        refs = [s for s in stock_data if s.get("Supplier_ID") == supplier_id]
        if refs:
            raise ValueError(
                f"Cannot delete: supplier is referenced by {len(refs)} stock item(s). "
                f"Delete those records first."
            )
        ws = wb["Suppliers"]
        data = _sheet_to_dicts(ws)
        data = [r for r in data if r.get("ID") != supplier_id]
        _dicts_to_sheet(ws, data, SHEETS["Suppliers"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


# ---- STOCK ----

@synchronized
def add_stock_item(item_name, category, quantity, purchase_price, selling_price,
                   min_quantity=5, supplier_id=None):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Stock"]
        sid = _next_id(ws)
        ws.append([sid, item_name, category, quantity, min_quantity,
                   purchase_price, selling_price, supplier_id, _now()])
        _save_and_close(wb)
        return sid
    except Exception:
        _close_wb(wb)
        raise


def get_stock_items(search="", category=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Stock"])
    wb.close()
    suppliers = {s["ID"]: s["Name"] for s in get_suppliers()}
    for r in data:
        r["supplier_name"] = suppliers.get(r.get("Supplier_ID"), "")
    if search:
        search = search.lower()
        data = [r for r in data if search in str(r.get("Item_Name", "")).lower()]
    if category and category != "All":
        data = [r for r in data if r.get("Category") == category]
    return data


def get_stock_item(stock_id):
    wb = _get_wb()
    for r in _sheet_to_dicts(wb["Stock"]):
        if r.get("ID") == stock_id:
            suppliers = {s["ID"]: s["Name"] for s in get_suppliers()}
            r["supplier_name"] = suppliers.get(r.get("Supplier_ID"), "")
            wb.close()
            return r
    wb.close()
    return None


@synchronized
def update_stock_item(stock_id, item_name, category, quantity, min_quantity,
                      purchase_price, selling_price, supplier_id=None):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Stock"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == stock_id:
                old_qty = r.get("Quantity", 0)
                r["Item_Name"] = item_name
                r["Category"] = category
                r["Quantity"] = quantity
                r["Min_Quantity"] = min_quantity
                r["Purchase_Price"] = purchase_price
                r["Selling_Price"] = selling_price
                r["Supplier_ID"] = supplier_id
                if old_qty != quantity:
                    diff = quantity - old_qty
                    _log_stock_change_internal(wb, stock_id, "adjustment", diff,
                                                old_qty, f"Manual: {old_qty} -> {quantity}")
                break
        _dicts_to_sheet(ws, data, SHEETS["Stock"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_stock_item(stock_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        sales_sheet = wb["Sales"] if "Sales" in wb.sheetnames else None
        sales_data = _sheet_to_dicts(sales_sheet) if sales_sheet else []
        refs = [s for s in sales_data if s.get("Stock_ID") == stock_id]
        if refs:
            raise ValueError(
                f"Cannot delete: stock item is referenced by {len(refs)} sale(s). "
                f"Delete those records first."
            )
        ws = wb["Stock"]
        data = _sheet_to_dicts(ws)
        data = [r for r in data if r.get("ID") != stock_id]
        _dicts_to_sheet(ws, data, SHEETS["Stock"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


def get_low_stock_items():
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Stock"])
    wb.close()
    return [r for r in data if r.get("Quantity", 0) <= r.get("Min_Quantity", 0)]


def get_categories():
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Stock"])
    wb.close()
    cats = set()
    for r in data:
        c = r.get("Category", "")
        if c:
            cats.add(c)
    return sorted(cats)


# ---- SALES ----

@synchronized
@synchronized
def record_sale(stock_id, quantity_sold, price, customer_id=None, 
                payment_status="paid", paid_amount=None, unpaid_amount=None):
    _check_write_lock()
    wb = _get_wb()
    try:
        stock_ws = wb["Stock"]
        stock_data = _sheet_to_dicts(stock_ws)
        item = None
        for r in stock_data:
            if r.get("ID") == stock_id:
                item = r
                break
        if not item:
            raise ValueError("Stock item not found")
        if item["Quantity"] < quantity_sold:
            raise ValueError(f'Not enough stock. Available: {item["Quantity"]}')
        total = round(quantity_sold * price, 2)

        if payment_status == "paid":
            paid_amt = total
            unpaid_amt = 0
        elif payment_status == "unpaid":
            paid_amt = 0
            unpaid_amt = total
        else:
            paid_amt = paid_amount if paid_amount is not None else 0
            unpaid_amt = unpaid_amount if unpaid_amount is not None else (total - paid_amt)

        sales_ws = wb["Sales"]
        sid = _next_id(sales_ws)
        receipt_no = f"RCP-{sid:06d}"
        payment_date = _now() if payment_status == "paid" else ""

        sales_ws.append([sid, stock_id, customer_id, quantity_sold, price, total,
                         payment_status, paid_amt, unpaid_amt, payment_date, _now(), receipt_no])
        old_qty = item["Quantity"]
        item["Quantity"] = old_qty - quantity_sold
        _dicts_to_sheet(stock_ws, stock_data, SHEETS["Stock"])
        _log_stock_change_internal(wb, stock_id, "sale", -quantity_sold,
                                    old_qty, f"Sold {quantity_sold} units")
        _save_and_close(wb)
        return sid, receipt_no
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def return_sale(sale_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        sales_ws = wb["Sales"]
        sales_data = _sheet_to_dicts(sales_ws)
        sale = None
        for r in sales_data:
            if r.get("ID") == sale_id:
                sale = r
                break
        if not sale:
            raise ValueError("Sale not found")
        stock_ws = wb["Stock"]
        stock_data = _sheet_to_dicts(stock_ws)
        for r in stock_data:
            if r.get("ID") == sale["Stock_ID"]:
                old_qty = r["Quantity"]
                r["Quantity"] = old_qty + sale["Quantity_Sold"]
                _log_stock_change_internal(wb, sale["Stock_ID"], "return",
                                            sale["Quantity_Sold"], old_qty,
                                            f"Return of sale #{sale_id}")
                break
        sales_data = [r for r in sales_data if r.get("ID") != sale_id]
        _dicts_to_sheet(sales_ws, sales_data, SHEETS["Sales"])
        _dicts_to_sheet(stock_ws, stock_data, SHEETS["Stock"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_sale(sale_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        sales_ws = wb["Sales"]
        sales_data = _sheet_to_dicts(sales_ws)
        sale = None
        for r in sales_data:
            if r.get("ID") == sale_id:
                sale = r
                break
        if sale:
            stock_ws = wb["Stock"]
            stock_data = _sheet_to_dicts(stock_ws)
            for r in stock_data:
                if r.get("ID") == sale["Stock_ID"]:
                    old_qty = r["Quantity"]
                    r["Quantity"] = old_qty + sale["Quantity_Sold"]
                    _log_stock_change_internal(wb, sale["Stock_ID"], "return",
                                                sale["Quantity_Sold"], old_qty,
                                                f"Deletion of sale #{sale_id}")
                    break
            _dicts_to_sheet(stock_ws, stock_data, SHEETS["Stock"])
        sales_data = [r for r in sales_data if r.get("ID") != sale_id]
        _dicts_to_sheet(sales_ws, sales_data, SHEETS["Sales"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def update_sale_payment(sale_id, payment_status, paid_amount=None, unpaid_amount=None):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Sales"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == sale_id:
                r["Payment_Status"] = payment_status
                total = _num(r.get("Total", 0))
                if payment_status == "paid":
                    r["Paid_Amount"] = total
                    r["Unpaid_Amount"] = 0
                    r["Payment_Date"] = _now()
                elif payment_status == "unpaid":
                    r["Paid_Amount"] = 0
                    r["Unpaid_Amount"] = total
                    r["Payment_Date"] = ""
                else:
                    r["Paid_Amount"] = paid_amount if paid_amount is not None else 0
                    r["Unpaid_Amount"] = unpaid_amount if unpaid_amount is not None else (total - (paid_amount or 0))
                    r["Payment_Date"] = _now() if (paid_amount or 0) > 0 else ""
                break
        _dicts_to_sheet(ws, data, SHEETS["Sales"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


def get_sales(search="", from_date="", to_date=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Sales"])
    wb.close()
    stock = {r["ID"]: r for r in get_stock_items()}
    customers = {r["ID"]: r for r in get_customers()}
    result = []
    for r in data:
        item = stock.get(r.get("Stock_ID"), {})
        customer = customers.get(r.get("Customer_ID"), {})
        r["item_name"] = item.get("Item_Name", "")
        r["category"] = item.get("Category", "")
        r["customer_name"] = customer.get("Name", "Walk-in Customer")
        r["customer_contact"] = customer.get("Contact", "")
        if search and search.lower() not in r["item_name"].lower() and search.lower() not in r["customer_name"].lower():
            continue
        sd = r.get("Sale_Date", "")
        if from_date and sd and sd[:10] < from_date:
            continue
        if to_date and sd and sd[:10] > to_date:
            continue
        result.append(r)
    result.sort(key=lambda x: x.get("Sale_Date") or "", reverse=True)
    return result


def get_daily_sales(date=None):
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    sales = get_sales()
    day_sales = [s for s in sales if (s.get("Sale_Date") or "")[:10] == date]
    extra = get_extra_income(from_date=date, to_date=date)
    extra_total = sum(_num(r.get("Amount", 0)) for r in extra)
    total_paid = sum(_num(s.get("Paid_Amount", 0)) for s in day_sales)
    total_unpaid = sum(_num(s.get("Unpaid_Amount", 0)) for s in day_sales)
    return {
        "count": len(day_sales),
        "total": sum(_num(s.get("Total", 0)) for s in day_sales),
        "total_paid": total_paid,
        "total_unpaid": total_unpaid,
        "extra_income": extra_total,
        "grand_total": sum(_num(s.get("Total", 0)) for s in day_sales) + extra_total,
    }


def get_monthly_report(year, month):
    wb = _get_wb()
    sales = _sheet_to_dicts(wb["Sales"])
    stock = {r["ID"]: r for r in _sheet_to_dicts(wb["Stock"])}
    wb.close()
    ym = f"{year}-{month:02d}"
    grouped = {}
    total_paid_month = 0
    total_unpaid_month = 0
    for s in sales:
        sd = str(s.get("Sale_Date") or "")
        if not sd.startswith(ym):
            continue
        item = stock.get(s.get("Stock_ID"), {})
        name = item.get("Item_Name", "Unknown")
        if name not in grouped:
            grouped[name] = {
                "item_name": name,
                "category": item.get("Category", ""),
                "qty_sold": 0,
                "price": s.get("Price", 0),
                "total_revenue": 0,
                "purchase_price": item.get("Purchase_Price", 0),
                "profit": 0,
            }
        qty = _num(s.get("Quantity_Sold", 0))
        price = _num(s.get("Price", 0))
        grouped[name]["qty_sold"] += qty
        grouped[name]["total_revenue"] += _num(s.get("Total", 0))
        grouped[name]["price"] = price
        total_paid_month += _num(s.get("Paid_Amount", 0))
        total_unpaid_month += _num(s.get("Unpaid_Amount", 0))
    extra = get_extra_income(from_date=f"{year}-{month:02d}-01",
                              to_date=f"{year}-{month:02d}-31")
    extra_total = sum(_num(r.get("Amount", 0)) for r in extra)
    result = []
    for name, g in grouped.items():
        cost = g["qty_sold"] * g["purchase_price"]
        g["profit"] = g["total_revenue"] - cost
        result.append(g)
    result.sort(key=lambda x: x["profit"], reverse=True)
    return {
        "items": result,
        "total_revenue": sum(g["total_revenue"] for g in result),
        "total_profit": sum(g["profit"] for g in result),
        "total_paid": total_paid_month,
        "total_unpaid": total_unpaid_month,
        "extra_income": extra_total,
        "net_income": sum(g["total_revenue"] for g in result) + extra_total,
    }


def get_yearly_report(year):
    sales = get_sales()
    monthly = {}
    for s in sales:
        sd = str(s.get("Sale_Date") or "")
        if not sd.startswith(str(year)):
            continue
        m = sd[5:7]
        if m not in monthly:
            monthly[m] = {"month": m, "sale_count": 0, "total_qty": 0,
                          "total_revenue": 0, "total_profit": 0,
                          "total_paid": 0, "total_unpaid": 0, "extra_income": 0}
        monthly[m]["sale_count"] += 1
        monthly[m]["total_qty"] += _num(s.get("Quantity_Sold", 0))
        monthly[m]["total_revenue"] += _num(s.get("Total", 0))
        monthly[m]["total_paid"] += _num(s.get("Paid_Amount", 0))
        monthly[m]["total_unpaid"] += _num(s.get("Unpaid_Amount", 0))
    stock = {r["ID"]: r for r in get_stock_items()}
    for s in sales:
        sd = str(s.get("Sale_Date") or "")
        if not sd.startswith(str(year)):
            continue
        m = sd[5:7]
        item = stock.get(s.get("Stock_ID"), {})
        cost = _num(s.get("Quantity_Sold", 0)) * _num(item.get("Purchase_Price", 0))
        monthly[m]["total_profit"] = _num(monthly[m].get("total_profit", 0)) + \
                                     _num(s.get("Total", 0)) - cost
    for m_key in monthly:
        ym = f"{year}-{m_key}"
        extra = get_extra_income(from_date=f"{ym}-01", to_date=f"{ym}-31")
        monthly[m_key]["extra_income"] = sum(_num(r.get("Amount", 0)) for r in extra)
    result = [monthly[m] for m in sorted(monthly.keys())]
    return result


def get_dashboard_stats():
    stock = get_stock_items()
    total_items = len(stock)
    low_stock = len([s for s in stock if _num(s.get("Quantity", 0)) <= _num(s.get("Min_Quantity", 0))])
    stock_value = sum(_num(s.get("Quantity", 0)) * _num(s.get("Purchase_Price", 0)) for s in stock)
    ds = get_daily_sales()
    total_sales = sum(_num(s.get("Total", 0)) for s in get_sales())
    total_extra = sum(_num(r.get("Amount", 0)) for r in get_extra_income())
    total_pending = sum(_num(s.get("Unpaid_Amount", 0)) for s in get_sales())
    return {
        "total_items": total_items,
        "low_stock": low_stock,
        "stock_value": round(stock_value, 2),
        "sales_today": round(ds.get("total", 0), 2),
        "extra_income_today": round(ds.get("extra_income", 0), 2),
        "total_revenue": round(total_sales + total_extra, 2),
        "total_pending": round(total_pending, 2),
    }


# ---- STOCK LOG ----

def _log_stock_change_internal(wb, stock_id, change_type, qty_change, old_qty, note=""):
    ws = wb["StockLog"]
    lid = _next_id(ws)
    new_qty = old_qty + qty_change
    ws.append([lid, stock_id, change_type, qty_change, old_qty, new_qty, note, _now()])


def get_stock_log(stock_id=None, limit=100):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["StockLog"])
    wb.close()
    if stock_id:
        data = [r for r in data if r.get("Stock_ID") == stock_id]
    data.sort(key=lambda x: x.get("Created_At") or "", reverse=True)
    return data[:limit]


def get_year_months():
    sales = get_sales()
    pairs = set()
    for s in sales:
        sd = str(s.get("Sale_Date") or "")
        if len(sd) >= 7:
            pairs.add((sd[:4], sd[5:7]))
    return sorted(pairs, reverse=True)


# ---- CUSTOMERS ----

@synchronized
def add_customer(name, contact="", address=""):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Customers"]
        cid = _next_id(ws)
        ws.append([cid, name, contact, address, _now()])
        _save_and_close(wb)
        return cid
    except Exception:
        _close_wb(wb)
        raise


def get_customers(search=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Customers"])
    wb.close()
    if search:
        search = search.lower()
        data = [r for r in data if search in str(r.get("Name", "")).lower()]
    return data


def get_customer(customer_id):
    wb = _get_wb()
    for r in _sheet_to_dicts(wb["Customers"]):
        if r.get("ID") == customer_id:
            wb.close()
            return r
    wb.close()
    return None


@synchronized
def update_customer(customer_id, name, contact, address):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Customers"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == customer_id:
                r["Name"] = name
                r["Contact"] = contact
                r["Address"] = address
                break
        _dicts_to_sheet(ws, data, SHEETS["Customers"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_customer(customer_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        sales_sheet = wb["Sales"] if "Sales" in wb.sheetnames else None
        preorders_sheet = wb["Preorders"] if "Preorders" in wb.sheetnames else None
        sales_data = _sheet_to_dicts(sales_sheet) if sales_sheet else []
        preorder_data = _sheet_to_dicts(preorders_sheet) if preorders_sheet else []
        sale_refs = [s for s in sales_data if s.get("Customer_ID") == customer_id]
        pre_refs = [p for p in preorder_data if p.get("Customer_ID") == customer_id]
        total_refs = len(sale_refs) + len(pre_refs)
        if total_refs > 0:
            details = []
            if sale_refs:
                details.append(f"{len(sale_refs)} sale(s)")
            if pre_refs:
                details.append(f"{len(pre_refs)} preorder(s)")
            raise ValueError(
                f"Cannot delete: customer is referenced by {' and '.join(details)}. "
                f"Delete those records first."
            )
        ws = wb["Customers"]
        data = _sheet_to_dicts(ws)
        data = [r for r in data if r.get("ID") != customer_id]
        _dicts_to_sheet(ws, data, SHEETS["Customers"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


# ---- PURCHASES (from suppliers) ----

@synchronized
def record_purchase(stock_id, supplier_id, quantity, cost_price):
    _check_write_lock()
    wb = _get_wb()
    try:
        stock_ws = wb["Stock"]
        stock_data = _sheet_to_dicts(stock_ws)
        item = None
        for r in stock_data:
            if r.get("ID") == stock_id:
                item = r
                break
        if not item:
            raise ValueError("Stock item not found")
        total_cost = round(quantity * cost_price, 2)
        old_qty = item["Quantity"]
        item["Quantity"] = old_qty + quantity
        item["Purchase_Price"] = cost_price
        _dicts_to_sheet(stock_ws, stock_data, SHEETS["Stock"])
        pur_ws = wb["Purchases"]
        pid = _next_id(pur_ws)
        pur_ws.append([pid, stock_id, supplier_id, quantity, cost_price, total_cost, _now()])
        _log_stock_change_internal(wb, stock_id, "purchase", quantity,
                                    old_qty, f"Purchased {quantity} units at \u20b9{cost_price}")
        _save_and_close(wb)
        return pid
    except Exception:
        _close_wb(wb)
        raise


def get_purchases(search=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Purchases"])
    wb.close()
    stock = {r["ID"]: r for r in get_stock_items()}
    suppliers = {s["ID"]: s["Name"] for s in get_suppliers()}
    result = []
    for r in data:
        r["item_name"] = stock.get(r.get("Stock_ID"), {}).get("Item_Name", "")
        r["supplier_name"] = suppliers.get(r.get("Supplier_ID"), "")
        if search and search.lower() not in r["item_name"].lower():
            continue
        result.append(r)
    return result


# ---- INVOICE FORMAT ----

def format_invoice_id(sale_id):
    try:
        from utils.company import load_company
        prefix = load_company().get("invoice_prefix", "INV")
    except (ImportError, AttributeError, OSError, json.JSONDecodeError, TypeError):
        prefix = "INV"
    return f"{prefix}-{sale_id:05d}"


# ---- EXTRA INCOME ----

@synchronized
def add_extra_income(source, description, amount, category, payment_method, reference_no=""):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["ExtraIncome"]
        sid = _next_id(ws)
        ws.append([sid, source, description, amount, category, payment_method, reference_no, _now()])
        _save_and_close(wb)
        return sid
    except Exception:
        _close_wb(wb)
        raise


def get_extra_income(search="", from_date="", to_date=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["ExtraIncome"])
    wb.close()
    result = []
    for r in data:
        if search and search.lower() not in str(r.get("Source", "")).lower():
            continue
        sd = r.get("Created_At", "")
        if from_date and sd and sd[:10] < from_date:
            continue
        if to_date and sd and sd[:10] > to_date:
            continue
        result.append(r)
    result.sort(key=lambda x: x.get("Created_At") or "", reverse=True)
    return result



@synchronized
def update_extra_income(income_id, source, description, amount, category, payment_method, reference_no):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["ExtraIncome"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == income_id:
                r["Source"] = source
                r["Description"] = description
                r["Amount"] = amount
                r["Category"] = category
                r["Payment_Method"] = payment_method
                r["Reference_No"] = reference_no
                break
        _dicts_to_sheet(ws, data, SHEETS["ExtraIncome"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_extra_income(income_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["ExtraIncome"]
        data = _sheet_to_dicts(ws)
        data = [r for r in data if r.get("ID") != income_id]
        _dicts_to_sheet(ws, data, SHEETS["ExtraIncome"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


def get_extra_income_summary(from_date="", to_date=""):
    incomes = get_extra_income(from_date=from_date, to_date=to_date)
    total = sum(_num(r.get("Amount", 0)) for r in incomes)
    by_category = {}
    by_method = {}
    for r in incomes:
        cat = r.get("Category", "Other")
        method = r.get("Payment_Method", "Cash")
        amt = _num(r.get("Amount", 0))
        by_category[cat] = by_category.get(cat, 0) + amt
        by_method[method] = by_method.get(method, 0) + amt
    return {
        "total": total,
        "count": len(incomes),
        "by_category": by_category,
        "by_method": by_method,
    }


# ---- PREORDERS ----

@synchronized
def add_preorder(customer_id, stock_id, quantity, preorder_price, delivery_date,
                 notes="", delivery_address="", advance_amount=0, advance_type="none"):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Preorders"]
        oid = _next_id(ws)
        total = round(quantity * preorder_price, 2)
        advance_paid_at = _now() if advance_amount > 0 else ""
        ws.append([oid, customer_id, stock_id, quantity, preorder_price,
                   total, delivery_date, notes, "pending",
                   _now(), "", "",
                   delivery_address, advance_amount, advance_type, advance_paid_at])
        _save_and_close(wb)
        return oid
    except Exception:
        _close_wb(wb)
        raise


def get_preorders(search="", status="", from_date="", to_date=""):
    wb = _get_wb()
    data = _sheet_to_dicts(wb["Preorders"])
    wb.close()
    stock = {r["ID"]: r for r in get_stock_items()}
    customers = {r["ID"]: r for r in get_customers()}
    result = []
    for r in data:
        item = stock.get(r.get("Stock_ID"), {})
        cust = customers.get(r.get("Customer_ID"), {})
        r["item_name"] = item.get("Item_Name", "")
        r["customer_name"] = cust.get("Name", "Unknown")
        if search and search.lower() not in r["customer_name"].lower() and search.lower() not in r["item_name"].lower():
            continue
        if status and r.get("Status", "pending") != status:
            continue
        sd = r.get("Created_At", "")
        if from_date and sd and sd[:10] < from_date:
            continue
        if to_date and sd and sd[:10] > to_date:
            continue
        result.append(r)
    result.sort(key=lambda x: x.get("Created_At") or "", reverse=True)
    return result



def get_preorder(preorder_id):
    for r in get_preorders():
        if r.get("ID") == preorder_id:
            return r
    return None


@synchronized
def update_preorder(preorder_id, customer_id, stock_id, quantity, preorder_price,
                    delivery_date, notes="", delivery_address="",
                    advance_amount=None, advance_type=None):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Preorders"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == preorder_id:
                r["Customer_ID"] = customer_id
                r["Stock_ID"] = stock_id
                r["Quantity"] = quantity
                r["Preorder_Price"] = preorder_price
                r["Total"] = round(quantity * preorder_price, 2)
                r["Delivery_Date"] = delivery_date
                r["Delivery_Address"] = delivery_address
                r["Notes"] = notes
                if advance_amount is not None:
                    r["Advance_Amount"] = advance_amount
                if advance_type is not None:
                    r["Advance_Payment_Type"] = advance_type
                    if advance_amount and advance_amount > 0:
                        r["Advance_Paid_At"] = _now()
                break
        _dicts_to_sheet(ws, data, SHEETS["Preorders"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def delete_preorder(preorder_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Preorders"]
        data = _sheet_to_dicts(ws)
        data = [r for r in data if r.get("ID") != preorder_id]
        _dicts_to_sheet(ws, data, SHEETS["Preorders"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def cancel_preorder(preorder_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Preorders"]
        data = _sheet_to_dicts(ws)
        for r in data:
            if r.get("ID") == preorder_id:
                r["Status"] = "cancelled"
                break
        _dicts_to_sheet(ws, data, SHEETS["Preorders"])
        _save_and_close(wb)
    except Exception:
        _close_wb(wb)
        raise


@synchronized
def complete_preorder(preorder_id):
    _check_write_lock()
    wb = _get_wb()
    try:
        ws = wb["Preorders"]
        data = _sheet_to_dicts(ws)
        preorder = None
        for r in data:
            if r.get("ID") == preorder_id:
                preorder = r
                break
        if not preorder:
            raise ValueError("Preorder not found")
        if preorder["Status"] != "pending":
            raise ValueError("Preorder is not in pending status")

        stock_id = preorder["Stock_ID"]
        qty = preorder["Quantity"]
        price = preorder["Preorder_Price"]
        customer_id = preorder.get("Customer_ID")

        stock_ws = wb["Stock"]
        stock_data = _sheet_to_dicts(stock_ws)
        item = None
        for r in stock_data:
            if r.get("ID") == stock_id:
                item = r
                break
        if not item:
            raise ValueError("Stock item not found")
        if item["Quantity"] < qty:
            raise ValueError(f'Not enough stock. Available: {item["Quantity"]}')

        total = round(qty * price, 2)
        advance_amt = _num(preorder.get("Advance_Amount", 0))
        advance_type = preorder.get("Advance_Payment_Type", "none")

        if advance_amt > 0 and advance_type == "full":
            payment_status = "paid"
            paid_amt = total
            unpaid_amt = 0
            payment_date = _now()
        elif advance_amt > 0 and advance_type == "partial":
            payment_status = "partial"
            paid_amt = advance_amt
            unpaid_amt = round(total - advance_amt, 2)
            payment_date = preorder.get("Advance_Paid_At", _now())
        else:
            payment_status = "unpaid"
            paid_amt = 0
            unpaid_amt = total
            payment_date = ""

        sales_ws = wb["Sales"]
        sid = _next_id(sales_ws)
        receipt_no = f"RCP-{sid:06d}"
        sales_ws.append([sid, stock_id, customer_id, qty, price, total,
                         payment_status, paid_amt, unpaid_amt, payment_date, _now(), receipt_no])

        old_qty = item["Quantity"]
        item["Quantity"] = old_qty - qty
        _dicts_to_sheet(stock_ws, stock_data, SHEETS["Stock"])
        _log_stock_change_internal(wb, stock_id, "sale", -qty,
                                    old_qty, f"Sold {qty} units (preorder #{preorder_id})")

        for r in data:
            if r.get("ID") == preorder_id:
                r["Status"] = "completed"
                r["Completed_At"] = _now()
                r["Sale_ID"] = sid
                break
        _dicts_to_sheet(ws, data, SHEETS["Preorders"])

        _save_and_close(wb)
        return sid, receipt_no
    except Exception:
        _close_wb(wb)
        raise
