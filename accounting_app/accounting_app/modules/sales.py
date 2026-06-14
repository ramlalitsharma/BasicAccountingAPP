import openpyxl
from openpyxl import Workbook
from tkinter import messagebox
import os

def record_sale(filepath, sale_data):
    """
    sale_data = [SaleID, ItemID, QuantitySold, Price, TotalAmount]
    """

    wb = openpyxl.load_workbook(filepath)

    # ✅ Ensure Stock sheet exists
    if "Stock" not in wb.sheetnames:
        messagebox.showerror("Error", "Stock sheet does not exist!")
        wb.close()
        return

    stock_ws = wb["Stock"]
    item_id, qty_sold = sale_data[1], sale_data[2]

    item_found = False
    for row in stock_ws.iter_rows(min_row=2):
        if str(row[0].value) == str(item_id):   # match Item ID
            item_found = True
            current_qty = row[3].value or 0

            if current_qty < qty_sold:
                messagebox.showerror("Error", f"Not enough stock for Item {item_id}. Available: {current_qty}")
                wb.close()
                return

            # Deduct quantity
            row[3].value = current_qty - qty_sold
            break

    if not item_found:
        messagebox.showerror("Error", f"Item {item_id} not available in stock!")
        wb.close()
        return

    # ✅ Ensure Sales sheet exists
    if "Sales" not in wb.sheetnames:
        sales_ws = wb.create_sheet("Sales")
        sales_ws.append(["Sale ID", "Item ID", "Quantity Sold", "Price", "Total Amount"])
    else:
        sales_ws = wb["Sales"]

    # Record sale in main workbook
    sales_ws.append(sale_data)
    wb.save(filepath)
    wb.close()

    # ✅ Create separate bill file on Desktop
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    bill_filename = os.path.join(desktop_path, f"Sale_{sale_data[0]}.xlsx")

    bill_wb = Workbook()
    bill_ws = bill_wb.active
    bill_ws.title = "Sale Bill"

    # Add headers
    bill_ws.append(["Sale ID", "Item ID", "Quantity Sold", "Price", "Total Amount"])
    # Add sale data
    bill_ws.append(sale_data)

    bill_wb.save(bill_filename)
    bill_wb.close()

    messagebox.showinfo("Success", f"Sale recorded and bill created on Desktop:\n{bill_filename}")


def get_sales(filepath):
    wb = openpyxl.load_workbook(filepath)

    # ✅ Create Sales sheet if missing
    if "Sales" not in wb.sheetnames:
        ws = wb.create_sheet("Sales")
        ws.append(["Sale ID", "Item ID", "Quantity Sold", "Price", "Total Amount"])
        wb.save(filepath)
        wb.close()
        return []  # no records yet

    ws = wb["Sales"]
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return data
