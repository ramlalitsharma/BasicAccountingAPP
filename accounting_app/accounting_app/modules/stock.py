import openpyxl
import os

def create_new_stock_file(filepath="data/transactions.xlsx"):
    if not os.path.exists("data"):
        os.makedirs("data")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["Item ID", "Item Name", "Category", "Quantity", "Purchase Price", "Selling Price"])
    wb.save(filepath)
    return filepath

def add_item(filepath, item_data):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Stock"]
    ws.append(item_data)
    wb.save(filepath)

def get_stock_items(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Stock"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]
