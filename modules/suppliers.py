import openpyxl
import os

def create_suppliers_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    if "Suppliers" not in wb.sheetnames:
        ws = wb.create_sheet("Suppliers")
        ws.append(["Supplier ID", "Name", "Contact", "Address", "Item Categories"])
        wb.save(filepath)

def add_supplier(filepath, supplier_data):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Suppliers"]
    ws.append(supplier_data)
    wb.save(filepath)

def get_suppliers(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Suppliers"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]
